import os
import re
from openpyxl import Workbook

# Directory to search
directory = "C:\\path\\to\\search\\directory"

# Output Excel file path
output_file = "C:\\path\\to\\output\\results.xlsx"

# List of search keys
search_keys = [
    "search_string_1",
    "search_string_2",
    "search_string_3"
]

# Dictionary to store results
results = {key: [] for key in search_keys}

# Recursively search files
for root, _, files in os.walk(directory):
    for file in files:
        file_path = os.path.join(root, file)
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line_number, line in enumerate(f, start=1):
                    # Check each search key
                    for key in search_keys:
                        if re.search(key, line, re.IGNORECASE):  # Ignore case
                            results[key].append({
                                "file": file_path,
                                "line_number": line_number,
                                "line": line.strip()
                            })
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")

wb = Workbook()

for key, entries in results.items():
    ws = wb.create_sheet(title=key[:30])  # Limit sheet name to 30 characters
    ws.append(["File Path", "Line Number", "Line Content"])
    for entry in entries:
        ws.append([entry["file"], entry["line_number"], entry["line"]])

# Remove the default empty sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

wb.save(output_file)
print(f"Results saved to {output_file}")